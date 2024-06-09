import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font

class ExcelViewerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Visor de Excel")
        
        # Frame principal
        main_frame = tk.Frame(root)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Frame izquierdo para los botones
        left_frame = tk.Frame(main_frame, bg='green')
        left_frame.pack(side=tk.LEFT, fill=tk.Y)

        # Frame central para el botón de generación de consolidado
        center_frame = tk.Frame(main_frame)
        center_frame.pack(side=tk.LEFT, padx=10, pady=10)

        # Frame derecho para los resultados
        right_frame = tk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Botón para cargar archivo 1
        self.load_button_1 = tk.Button(left_frame, text="Adjuntar Archivo EXCEL - 1", command=lambda: self.load_file(1), bg='green', fg='white', wraplength=150)
        self.load_button_1.pack(pady=10, padx=10)

        # Botón para cargar archivo 2
        self.load_button_2 = tk.Button(left_frame, text="Adjuntar Archivo EXCEL - 2", command=lambda: self.load_file(2), bg='green', fg='white', wraplength=150)
        self.load_button_2.pack(pady=10, padx=10)

        # Botón para generar archivo
        self.combine_button = tk.Button(center_frame, text="Generar Archivo", command=self.combine_columns, bg='green', fg='white')
        self.combine_button.pack(pady=10)

        # Área de texto para mostrar datos combinados
        self.text_area = tk.Text(right_frame)
        self.text_area.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.df1 = None
        self.df2 = None

        # Combobox y entrada de nombre de columna para la configuración de columnas
        self.column_combobox_1 = ttk.Combobox(left_frame, state="readonly")
        self.column_combobox_1.pack(pady=10)
        
        self.column_combobox_2 = ttk.Combobox(left_frame, state="readonly")
        self.column_combobox_2.pack(pady=10)
        
        self.new_column_name_entry = tk.Entry(left_frame)
        self.new_column_name_entry.pack(pady=10)
        self.new_column_name_entry.insert(0, "Nombre de Nueva Columna")

    def load_file(self, file_number):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            return
        
        try:
            # Lee el archivo Excel usando la primera fila como encabezado
            df = pd.read_excel(file_path, header=0)
            if file_number == 1:
                self.df1 = df
                self.file_path_1 = file_path
                self.column_combobox_1['values'] = list(self.df1.columns)
                messagebox.showinfo("Éxito", "Archivo 1 cargado correctamente")
            elif file_number == 2:
                self.df2 = df
                self.file_path_2 = file_path
                self.column_combobox_2['values'] = list(self.df2.columns)
                messagebox.showinfo("Éxito", "Archivo 2 cargado correctamente")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar el archivo: {e}")

    def combine_columns(self):
        column_1 = self.column_combobox_1.get()
        column_2 = self.column_combobox_2.get()
        new_column_name = self.new_column_name_entry.get()
        
        if self.df1 is None or self.df2 is None:
            messagebox.showerror("Error", "Debe cargar ambos archivos Excel primero.")
            return
        
        if not column_1 or not column_2 or not new_column_name:
            messagebox.showerror("Error", "Debe seleccionar ambas columnas y proporcionar un nombre para la nueva columna.")
            return
        
        # Pedir al usuario que seleccione el archivo de destino
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx *.xls")])
        if not save_path:
            return
        
        # Combina las columnas y guarda en un nuevo archivo manteniendo los estilos
        try:
            wb1 = load_workbook(self.file_path_1)
            wb2 = load_workbook(self.file_path_2)
            
            ws1 = wb1.active
            ws2 = wb2.active
            
            combined_data = pd.concat([self.df1[column_1], self.df2[column_2]], ignore_index=True)
            combined_df = pd.DataFrame({new_column_name: combined_data})
            
            # Crear un nuevo libro y hoja para guardar el archivo combinado
            new_wb = load_workbook(self.file_path_1)  # Cargar el archivo base para mantener estilos
            new_ws = new_wb.create_sheet(title="Consolidado")
            
            for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    new_cell = new_ws.cell(row=r_idx+1, column=c_idx+1, value=value)
                    if r_idx == 0:
                        # Aplicar estilo de encabezado
                        new_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        new_cell.font = Font(bold=True)
            
            new_wb.save(save_path)
            self.text_area.delete(1.0, tk.END)
            self.text_area.insert(tk.END, combined_df.to_string(index=False))
            messagebox.showinfo("Éxito", f"Archivo guardado correctamente en {save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo combinar las columnas: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelViewerApp(root)
    root.geometry("1200x600")
    root.mainloop()
